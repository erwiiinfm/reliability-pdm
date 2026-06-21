from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from typing import Optional
import pandas as pd
import io
import os
from pathlib import Path

router = APIRouter(prefix='/api/assets', tags=['Asset Management'])

_candidate = Path(__file__).parents[2] / '1. Asset Management.xlsx'
ASSET_FILE = _candidate if _candidate.exists() else Path(__file__).parents[3] / '1. Asset Management.xlsx'

# In-memory store — load once, edits kept in memory
_assets: dict[str, dict] = {}  # keyed by Equipment Register

COLS = [
    'Equipment Register', 'Model', 'Tipe APPT', 'Tipe', 'Kelas',
    'Tahun Perolehan', 'Unit/ Chassis', 'Engine', 'Lokasi Operasi',
    'Divisi', 'Departemen', 'PIC Perawatan', 'Dinas', 'Kontrak',
    'Status Aset', 'Nomor Aset', 'WO', 'Status Lapor Rawatan'
]


def load_assets():
    global _assets
    if _assets:
        return
    df = pd.read_excel(str(ASSET_FILE), sheet_name='APPT Profile', header=4)
    for _, row in df.iterrows():
        code = str(row.get('Equipment Register', '')).strip()
        if not code or code == 'nan':
            continue
        _assets[code] = {
            c: ('' if pd.isna(row.get(c)) else str(row.get(c, '')).strip())
            for c in COLS
        }


load_assets()


class AssetUpdate(BaseModel):
    Model: Optional[str] = None
    Tipe_APPT: Optional[str] = None
    Tipe: Optional[str] = None
    Kelas: Optional[str] = None
    Tahun_Perolehan: Optional[str] = None
    Unit_Chassis: Optional[str] = None
    Engine: Optional[str] = None
    Lokasi_Operasi: Optional[str] = None
    Divisi: Optional[str] = None
    Departemen: Optional[str] = None
    PIC_Perawatan: Optional[str] = None
    Dinas: Optional[str] = None
    Kontrak: Optional[str] = None
    Status_Aset: Optional[str] = None
    Nomor_Aset: Optional[str] = None
    WO: Optional[str] = None
    Status_Lapor_Rawatan: Optional[str] = None


# Map Pydantic field names → Excel column names
_FIELD_MAP = {
    'Model': 'Model',
    'Tipe_APPT': 'Tipe APPT',
    'Tipe': 'Tipe',
    'Kelas': 'Kelas',
    'Tahun_Perolehan': 'Tahun Perolehan',
    'Unit_Chassis': 'Unit/ Chassis',
    'Engine': 'Engine',
    'Lokasi_Operasi': 'Lokasi Operasi',
    'Divisi': 'Divisi',
    'Departemen': 'Departemen',
    'PIC_Perawatan': 'PIC Perawatan',
    'Dinas': 'Dinas',
    'Kontrak': 'Kontrak',
    'Status_Aset': 'Status Aset',
    'Nomor_Aset': 'Nomor Aset',
    'WO': 'WO',
    'Status_Lapor_Rawatan': 'Status Lapor Rawatan',
}


@router.get('')
def list_assets(
    q: Optional[str] = Query(None, description='Search query'),
    tipe: Optional[str] = Query(None, description='Filter by Tipe APPT'),
    status: Optional[str] = Query(None, description='Filter by Status Aset'),
    lokasi: Optional[str] = Query(None, description='Filter by Lokasi Operasi'),
    limit: int = Query(100, ge=1, le=5000),
    offset: int = Query(0, ge=0),
):
    load_assets()
    results = list(_assets.values())

    if q:
        q_lower = q.lower()
        search_cols = ['Equipment Register', 'Model', 'Tipe', 'Engine', 'Unit/ Chassis']
        results = [
            r for r in results
            if any(q_lower in r.get(c, '').lower() for c in search_cols)
        ]
    if tipe:
        results = [r for r in results if r.get('Tipe APPT', '') == tipe]
    if status:
        results = [r for r in results if r.get('Status Aset', '') == status]
    if lokasi:
        results = [r for r in results if r.get('Lokasi Operasi', '') == lokasi]

    total = len(results)
    page = results[offset: offset + limit]
    return {'total': total, 'offset': offset, 'limit': limit, 'data': page}


@router.get('/meta/filters')
def get_filters():
    load_assets()
    all_assets = list(_assets.values())

    def unique_sorted(col):
        return sorted({r.get(col, '') for r in all_assets if r.get(col, '')})

    return {
        'tipe': unique_sorted('Tipe APPT'),
        'status': unique_sorted('Status Aset'),
        'lokasi': unique_sorted('Lokasi Operasi'),
        'model': unique_sorted('Model'),
    }


@router.get('/export/csv')
def export_csv():
    load_assets()
    df = pd.DataFrame(list(_assets.values()), columns=COLS)
    output = io.StringIO()
    df.to_csv(output, index=False)
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type='text/csv',
        headers={'Content-Disposition': 'attachment; filename="assets_ptba.csv"'},
    )


@router.get('/{unit_code}')
def get_asset(unit_code: str):
    load_assets()
    asset = _assets.get(unit_code)
    if not asset:
        raise HTTPException(status_code=404, detail=f'Asset {unit_code} not found')
    return asset


@router.put('/{unit_code}')
def update_asset(unit_code: str, payload: AssetUpdate):
    load_assets()
    if unit_code not in _assets:
        raise HTTPException(status_code=404, detail=f'Asset {unit_code} not found')

    updates = payload.model_dump(exclude_none=True)
    col_updates: dict[str, str] = {}
    for field, value in updates.items():
        col_name = _FIELD_MAP.get(field)
        if col_name:
            _assets[unit_code][col_name] = value
            col_updates[col_name] = value

    # Try to persist back to xlsx
    try:
        import openpyxl
        wb = openpyxl.load_workbook(str(ASSET_FILE))
        ws = wb['APPT Profile']

        # Header row is row index 4 (0-based) → Excel row 5
        # Find column indices from header row (row 5 in 1-based)
        header_row = 5
        col_idx: dict[str, int] = {}
        for cell in ws[header_row]:
            if cell.value and str(cell.value).strip() in col_updates:
                col_idx[str(cell.value).strip()] = cell.column

        # Find the row with matching Equipment Register
        eq_col = None
        for cell in ws[header_row]:
            if cell.value and str(cell.value).strip() == 'Equipment Register':
                eq_col = cell.column
                break

        if eq_col:
            for row_num in range(header_row + 1, ws.max_row + 1):
                cell_val = ws.cell(row=row_num, column=eq_col).value
                if cell_val and str(cell_val).strip() == unit_code:
                    for col_name, value in col_updates.items():
                        if col_name in col_idx:
                            ws.cell(row=row_num, column=col_idx[col_name]).value = value
                    break

        wb.save(str(ASSET_FILE))
    except Exception:
        pass  # Memory already updated; xlsx write failure is non-fatal

    return {'status': 'ok', 'unit_code': unit_code, 'updated': col_updates}
